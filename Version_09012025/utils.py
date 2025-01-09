import ultralytics
from tqdm import tqdm
from pdf2image import convert_from_path
from ultralytics import YOLO
import shutil
import glob
import spacy
import os
import pdfplumber
import re
import pandas as pd
import numpy as np
from PIL import Image
import cv2
from google.colab.patches import cv2_imshow

from typing_extensions import ParamSpecKwargs
import camelot
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple


def create_classified_folder(classified_folder):
    classes = ['appendix', 'cover', 'description', 'description and table',
              'graph', 'info', 'list figure', 'list table', 'table', 'table content'
              ]
    for c in classes:
        os.makedirs(classified_folder + '/' + c, exist_ok=True)

# Function to add space between concatenated words
def add_space_to_text(text):
    return re.sub(r'([a-z])([A-Z])', r'\1 \2', text)

def clean_unit(value):
    # If value is a float or int, return it as is
    if isinstance(value, (int, float)):
        return value
    # If value is a string, use regex to extract the number
    if isinstance(value, str):
        match = re.search(r'\d+\.?\d*', value)
        if match:
            return float(match.group())
        else:
            return np.nan  # Return None if no number is found
    # If value is neither a float, int, nor string, return None
    return np.nan

def camelot_detect_table(report, page, flavor="stream"):
    # Extract tables
    page = str(page)
    tables = camelot.read_pdf(report, pages=page, flavor="stream")

    # Because it returns multiple tables e.g. tables[0], tables[1]
    # Select the largest table to avoid picking always tables[0]
    if len(tables) == 1:
        # If there's only one table, use it directly
        largest_table = tables[0]
    else:
        # If there are multiple tables, select the largest one
        largest_table = max(tables, key=lambda t: t.df.shape[0] * t.df.shape[1])

    # Convert to DataFrame
    df = largest_table.df
    return df

def query_to_page(df_table, query, sample):
    # Define mapping for query conditions
    query_conditions = {
        'Compositional data - Recombined Fluid': {
            'ANALYSIS': 'COMPOSITIONAL ANALYSIS',
            'SAMPLE': sample,
            'FLUID': 'RESERVOIR FLUID'
        },
        'Compositional data - Separator Fluid': {
            'ANALYSIS': 'COMPOSITIONAL ANALYSIS',
            'SAMPLE': sample,
            'FLUID': 'SEPARATOR FLUID'
        },
        'Compositional data - Flashed Oil': {
            'ANALYSIS': 'COMPOSITIONAL ANALYSIS',
            'SAMPLE': sample,
            'FLUID': 'FLASHED OIL'
        },
        'Compositional data - Flashed Gas': {
            'ANALYSIS': 'COMPOSITIONAL ANALYSIS',
            'SAMPLE': sample,
            'FLUID': 'FLASHED GAS'
        },
        'CCE - Recombined Fluid': {
            'EXPERIMENT': 'CONSTANT COMPOSITION EXPANSION', # for some cases can be EXPERIMENT or ANALYSIS
            'SAMPLE': sample
        },
        'CVD - Fluid Recovery': {
            'EXPERIMENT': ['CONSTANT VOLUME DEPLETION', 'CVD'],
            'ANALYSIS': 'FLUID RECOVERY',
            'SAMPLE': sample
        },
        'CVD - Wellstream Properties': {
            'EXPERIMENT': ['CONSTANT VOLUME DEPLETION', 'CVD'],
            'ANALYSIS': 'PRODUCED WELLSTREAM PROPERTIES',
            'SAMPLE': sample
        },
        'CVD - Wellstream Compositions': {
            'EXPERIMENT': ['CONSTANT VOLUME DEPLETION', 'CVD'],
            'ANALYSIS': ['PRODUCED WELLSTREAM COMPOSITIONAL ANALYSIS', 'WELLSTREAM COMPOSITIONS'],
            'SAMPLE': sample
        }
    }

    # Retrieve conditions for the selected query
    conditions = query_conditions.get(query, {})

    # Filter the DataFrame based on conditions
    if conditions:
        df_query = df_table
        for column, value in conditions.items():
            if isinstance(value, list):  # Handle multiple values for a column
                df_query = df_query[df_query[column].apply(lambda x: any(val in str(x) for val in value))]
            else:
                # Allow substring match and handle repeated terms like 'CONSTANT VOLUME DEPLETION, CONSTANT VOLUME DEPLETION'
                if column == 'ANALYSIS' or column == 'EXPERIMENT':
                    df_query = df_query[df_query[column].str.contains(value, na=False)]
                else:
                    df_query = df_query[df_query[column] == value]
    else:
        df_query = pd.DataFrame()  # Empty DataFrame if query is not found

    # Get page number
    query_page = df_query['page_no'].values

    return query_page


from openpyxl import load_workbook

def write_info_to_excel(sample, df_info, output_path):
    # Filter the DataFrame based on the sample type
    clean_sample = sample.replace('SAMPLE', '').strip()

    if clean_sample=='RECOMBINED':
      df_info_sample = df_info[df_info['Sample Type'] == clean_sample]
    else:
      df_info_sample = df_info[df_info['Sample Name'] == clean_sample]


    # Load the existing Excel workbook
    sample_filename = output_path + sample + '.xlsx'
    wb = load_workbook(sample_filename)

    # Define mapping of columns to sheets and starting cells
    column_sheet_map = {
        'Sample Name': [('General Info', 'B11')],
        'Sample Type': [('General Info', 'B13')],
        'Sample Date': [('General Info', 'B15')],
        'Well Name': [('General Info', 'B17')],
        'Reservoir': [('General Info', 'B19')],
        'Formation': [('General Info', 'B21')],
        'Sampling Pressure': [('General Info', 'G25')],
        'Sampling Temperature': [('General Info', 'G27')],
        'Reservoir Pressure': [
            ('CVD Experimental Data', 'E6'),
            ('CCE Experimental Data', 'E5')  # Add second target
        ],
        'Reservoir Temperature': [
            ('CVD Experimental Data', 'E7'),
            ('CCE Experimental Data', 'E6')  # Add second target
        ],
        'Flash Oil Density': [('Compositional Data', 'N4')] # the only info added to this sheet
    }

    # Iterate through columns and write them to respective sheets and cells
    for column, targets in column_sheet_map.items():
        for sheet_name, start_cell in targets:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                print(f"Sheet '{sheet_name}' does not exist in the template.")
                continue

            # Write the first value from the filtered DataFrame for the column
            if column in df_info_sample.columns and not df_info_sample[column].empty:
                value = df_info_sample[column].iloc[0]
                start_row = int(start_cell[1:])
                start_col = ord(start_cell[0].upper()) - ord('A') + 1
                sheet.cell(row=start_row, column=start_col, value=value)

    # Save the updated workbook
    wb.save(sample_filename)
    print(f"Info successfully written to {sample_filename}.")

def write_list_to_excel(workbook, values, sheet_name, start_cell="A1"):
    """
    Write a list of values into consecutive cells in a specified sheet and row of an Excel file.

    :param file_path: Path to the Excel file. If it doesn't exist, a new one will be created.
    :param sheet_name: Name of the sheet where data will be written.
    :param values: List of values to write.
    :param start_cell: Starting cell (e.g., 'A1') for writing the values.
    """

    sheet = workbook[sheet_name]

    # Convert start_cell to row and column indices
    start_row, start_col = coordinate_to_tuple(start_cell)

    # Write values into consecutive cells
    for col_offset, value in enumerate(values):
        sheet.cell(row=start_row, column=start_col + col_offset, value=value)


# Parse Starting Cell
def parse_start_cell(start_cell):
    start_col = ord(start_cell[0].upper()) - ord('A') + 1
    start_row = int(start_cell[1:])
    return start_row, start_col

# Write DataFrame to Excel
def write_dataframe_to_excel(df, workbook, sheet_name, start_cell):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

    sheet = workbook[sheet_name]
    start_row, start_col = parse_start_cell(start_cell)

    for i, row in enumerate(df.itertuples(index=False), start=start_row):
        for j, value in enumerate(row, start=start_col):
            cell = sheet.cell(row=i, column=j)
            if cell.value is None or cell.value == '':
                cell.value = value
            else:
                print(f"Cell ({i}, {j}) in sheet '{sheet_name}' already contains a value. Skipping overwrite.")

def write_value_to_excel(workbook, sheet_name, cell, value):
    """
    Write a single value to a specific cell in the Excel workbook, without overwriting existing values.
    :param workbook: Loaded openpyxl Workbook object.
    :param sheet_name: Name of the sheet where the value will be written.
    :param cell: Cell address (e.g., 'I3') where the value will be written.
    :param value: The value to write.
    """
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")
    sheet = workbook[sheet_name]
    row, col = parse_start_cell(cell)

    # Check if the cell already has a value
    current_value = sheet.cell(row=row, column=col).value
    if current_value is None or current_value == '':
        sheet.cell(row=row, column=col, value=value)
    else:
        # print(f"Cell {cell} in sheet '{sheet_name}' already contains a value. Skipping overwrite.")
        pass
        
# Main Process
def process_pdf_to_excel(df_table, sample, report, template_path, output_path, all_queries):
    workbook = load_workbook(template_path)

    pages = []
    for query in all_queries:
        query_page = query_to_page(df_table, query, sample).tolist()
        if not query_page:
            # print(f'Page of Query {query} NOT FOUND')
            continue

        # print(f'Page of Query {query} FOUND: Page {query_page}')

        pvt_table = None  # Initialize pvt_table for the current query

        multitables = []
        for p in query_page:
            # Read table from the PDF using Camelot
            tables = camelot_detect_table(report, p)

            # Apply slicing logic
            if query == 'Compositional data - Recombined Fluid':
                # Get molecular weight and GOR'
                try:
                    molecular_weight = tables.iloc[4,6]
                    recombine_gor = tables.iloc[34,6]
                    c30_mw = tables.iloc[29,6]

                    # Write these variables to Excel
                    if query in variable_config:
                        if 'molecular_weight' in variable_config[query]:
                            write_value_to_excel(workbook,
                                                 variable_config[query]['molecular_weight']['sheet'],
                                                 variable_config[query]['molecular_weight']['cell'],
                                                 molecular_weight)
                        if 'recombine_gor' in variable_config[query]:
                            write_value_to_excel(workbook,
                                                 variable_config[query]['recombine_gor']['sheet'],
                                                 variable_config[query]['recombine_gor']['cell'],
                                                 recombine_gor)
                        if 'c30_mw' in variable_config[query]:
                            write_value_to_excel(workbook,
                                                 variable_config[query]['c30_mw']['sheet'],
                                                 variable_config[query]['c30_mw']['cell'],
                                                 c30_mw)

                except:
                    # print('ERROR writing Molecular Weight and GOR')
                    pass

                tables = tables.iloc[2:-1, :5]
                colname = ['Boiling Point (K)', 'Component', 'Formula', 'Mole Amounts', 'Mass Amounts']

            if query == 'Compositional data - Separator Fluid':
                # Get molecular weight and GOR
                try:
                    molecular_weight = tables.iloc[4,6]
                    recombine_gor = tables.iloc[34,6]

                    # Write these variables to Excel
                    if query in variable_config:
                        if 'molecular_weight' in variable_config[query]:
                            write_value_to_excel(workbook,
                                                 variable_config[query]['molecular_weight']['sheet'],
                                                 variable_config[query]['molecular_weight']['cell'],
                                                 molecular_weight)
                        if 'recombine_gor' in variable_config[query]:
                            write_value_to_excel(workbook,
                                                 variable_config[query]['recombine_gor']['sheet'],
                                                 variable_config[query]['recombine_gor']['cell'],
                                                 recombine_gor)
                        if 'c30_mw' in variable_config[query]:
                            write_value_to_excel(workbook,
                                                 variable_config[query]['c30_mw']['sheet'],
                                                 variable_config[query]['c30_mw']['cell'],
                                                 c30_mw)

                except:
                    # print('ERROR writing Molecular Weight and GOR')
                    pass

                tables = tables.iloc[2:-1, :5]
                colname = ['Boiling Point (K)', 'Component', 'Formula', 'Mole Amounts', 'Mass Amounts']

            elif query == 'Compositional data - Flashed Oil':
                try:
                    # Get molecular weight and oil density
                    molecular_weight = tables.iloc[4,6]
                    fluid_density = tables.iloc[36,6]

                    # Write these variables to Excel
                    if query in variable_config:
                        if 'molecular_weight' in variable_config[query]:
                            write_value_to_excel(workbook,
                                                 variable_config[query]['molecular_weight']['sheet'],
                                                 variable_config[query]['molecular_weight']['cell'],
                                                 molecular_weight)
                        if 'fluid_density' in variable_config[query]:
                            write_value_to_excel(workbook,
                                                 variable_config[query]['fluid_density']['sheet'],
                                                 variable_config[query]['fluid_density']['cell'],
                                                 fluid_density)
                        # if 'c30_mw' in variable_config[query]:
                        #     write_value_to_excel(workbook,
                        #                          variable_config[query]['c30_mw']['sheet'],
                        #                          variable_config[query]['c30_mw']['cell'],
                        #                          c30_mw)

                except:
                    # print('ERROR writing Molecular Weight and Fluid Density')
                    pass

                tables = tables.iloc[2:-1, :5]
                colname = ['Boiling Point (K)', 'Component', 'Formula', 'Mole Amounts', 'Mass Amounts']

            elif query == 'Compositional data - Flashed Gas':
                try:
                    # Get molecular weight and specific gravity
                    molecular_weight = clean_unit(tables.iloc[24,2])
                    specific_gravity = clean_unit(tables.iloc[25,2])

                    # Write these variables to Excel
                    if query in variable_config:
                        if 'molecular_weight' in variable_config[query]:
                            write_value_to_excel(workbook,
                                                 variable_config[query]['molecular_weight']['sheet'],
                                                 variable_config[query]['molecular_weight']['cell'],
                                                 molecular_weight)
                        if 'specific_gravity' in variable_config[query]:
                            write_value_to_excel(workbook,
                                                 variable_config[query]['specific_gravity']['sheet'],
                                                 variable_config[query]['specific_gravity']['cell'],
                                                 specific_gravity)
                except:
                    # print('ERROR Writing Molecular Weight and Specific Gravity')
                    pass

                tables = tables.iloc[2:-8, :4]
                colname = ['Component', 'Formula', 'Mole Amounts', 'Mass Amounts']

            elif query == 'CCE - Recombined Fluid':
                tables = tables.iloc[3:, :]
                colname = ['Pressure (psia)', 'Pressure (MPa)', 'Relative Volume', 'Z-Factor', 'Y-Function', 'Liquid Volume', 'Fluid Density']

            elif query == 'CVD - Fluid Recovery':
                tables = tables.iloc[4:, :]
                colname = ['Pressure (psia)', 'Pressure (MPa)', 'Liquid Drop Out', 'Cum Produced Fluid', 'Cum Liquid Recovery in STB/MMscf', 'Cum Liquid Recovery in m3/10^6m', 'Separator Condensate GOR in STB/MMscf', 'Separator Condensate GOR in m3/10^6m']

            elif query == 'CVD - Wellstream Properties':
                tables = tables.iloc[5:, :]
                colname = ['Pressure (psia)', 'Pressure (MPa)', 'Gas Density', 'Gas Viscosity', 'Gas Deviation Factor', 'Two-Phase Gas Deviation Factor', 'p/z (psia)', 'p/z (MPa)', 'p/z_2ph (psia)', 'p/z_2ph (MPa)']

            elif query == 'CVD - Wellstream Compositions':
                colname = ['Name', 'Formula'] + list(tables.iloc[2:3, 2:].values[0])
                tables = tables.iloc[4:, :]

            # Clean tables and rename columns
            tables = tables.replace('', np.nan).dropna(axis=0, how='all').dropna(axis=1, how='all')
            tables.columns = colname

            # Add additional tables
            tables['Reported MW'] = np.nan

            # Form `pvt_table` after cleaning
            if query == 'Compositional data - Recombined Fluid':
                pvt_table = tables[['Formula', 'Reported MW', 'Mole Amounts', 'Mass Amounts']]
            if query == 'Compositional data - Separator Fluid':
                pvt_table = tables[['Formula', 'Reported MW', 'Mole Amounts', 'Mass Amounts']]
            elif query == 'Compositional data - Flashed Oil':
                pvt_table = tables[['Formula', 'Reported MW', 'Mole Amounts', 'Mass Amounts']]
            elif query == 'Compositional data - Flashed Gas':
                pvt_table = tables[['Formula', 'Reported MW', 'Mole Amounts', 'Mass Amounts']]
            elif query == 'CCE - Recombined Fluid':
                pvt_table = tables[['Pressure (psia)', 'Relative Volume', 'Fluid Density', 'Liquid Volume', 'Z-Factor']]
            elif query == 'CVD - Fluid Recovery':
                pvt_table = tables[['Pressure (psia)', 'Liquid Drop Out', 'Cum Produced Fluid']]
            elif query == 'CVD - Wellstream Properties':
                pvt_table = tables[['Gas Deviation Factor', 'Gas Density', 'Gas Viscosity']]
            elif query == 'CVD - Wellstream Compositions':
                pvt_table = tables.drop(columns=['Formula'])  # Needs further modification later
                # pvt_table = tables.copy()  # Needs further modification later

            # print(tables.head(), '\n')
            multitables.append(pvt_table)

        # Special condition For CVD Wellstream Compositions
        if query == 'CVD - Wellstream Compositions':
            # Merge both dataframes
            pvt_table = pd.merge(multitables[0], multitables[1], on=['Name', 'Reported MW'], how='inner')

            # Rearrange column Reported MW
            col = pvt_table.pop('Reported MW')
            pvt_table.insert(1, 'Reported MW', col)

            # Write the pressures of CVD as column name
            pressure_col = pvt_table.columns[2:]
            pressure_col = [str(i)+str(' psia') for i in pressure_col]
            # print(pressure_col)
            write_list_to_excel(workbook, pressure_col,
                                sheet_name="CVD Experimental Data",
                                start_cell="Y4")


        # Write to Excel if query is mapped
        if query in config and pvt_table is not None:
            sheet_name = config[query]['sheet']
            start_cell = config[query]['start_cell']
            write_dataframe_to_excel(pvt_table, workbook, sheet_name, start_cell)
        else:
            # print(f"No mapping found for query '{query}' in config or no data extracted for this query.")
            pass
            
    pages.append(query_page)

    # Save the updated workbook
    sample_filaname = output_path + sample + '.xlsx'
    workbook.save(sample_filaname)
    print(f"Data successfully written to {sample_filaname}.")
